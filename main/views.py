import csv
from datetime import datetime
import itertools
import json
import math
import re
from unicodedata import category
import xml.etree.ElementTree as ET

if __name__ != '__main__':
    from django.conf import settings
    from django.http import HttpResponse, JsonResponse
    from django.shortcuts import render
    from django.template.loader import render_to_string

    STATIC_DIR = settings.BASE_DIR / 'main' / 'static' / 'main'
else:
    from pathlib import Path
    BASE_DIR = Path(__file__).resolve().parent.parent
    STATIC_DIR = BASE_DIR / 'main' / 'static' / 'main'

from pyhtml import *

# from main.models import RFK

RFK_REGEX = r"[bdes](?:\d{3,}\.\d+)"
SCORE_CLASSES = ['', 'w3-pale-yellow', 'w3-yellow', 'w3-pale-red', 'w3-red']
SCORE_SCALE = [5, 25, 50, 95, 100]
ROWS, COLUMNS, IGNORE = (('d'), ('b'), ('s', 'e')) # Milliseid koodigruppe ja kus arvesse v6tta arvutamisel
COLUMNS_VERBOSE, IGNORE_VERBOSE = (('b', 's'), ('e')) # Milliseid koodigruppe ja kus arvesse v6tta verbaalsel esitusel
LEVEL_MILD, LEVEL_MODERATE, LEVEL_SEVERE, LEVEL_EXTREME = range(1, 5) # RFK määrajad
LEVEL_TTA = 9 # RFK määraja TTa e Täpsustama
FILE_VER = 'icf2017_est_2022_v2'

"""
Terminoloogia:
Level of functioning                Parallel level of disability
Body functions and structures       Impairments
Activities                          Activity limitations
Participation                       Participation restrictions

polytomous (multiple-level) scale:
“mild”, “moderate”, “severe” or “extreme"
"""

class ICF_Est_New():
    # Loeb ICF kooditabeli sõnastikuks:
    # keys = code
    # element.keys:
    # 'version',
    # 'Dimension', 'Chapter',
    # 'Block', 'SecondLevel', 'ThirdLevel', 'FourthLevel', 'levelno',
    # 'code', 'parent', 'mlsort', 'leafnode',
    # 'Title', 'Description', 'Inclusions', 'Exclusions', 'selected',
    # 'Translated_title', 'Translated_description', 'Translated_inclusions', 'Translated_exclusions']

    def __init__(self):
        self.df = dict()
        with open(STATIC_DIR / f'{FILE_VER}.csv', newline='', encoding='windows-1252') as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';', quotechar='"')
            n = 0
            for row in reader:
                if row['code']:
                    self.df[row['code']] = row
                    n += 1
        # print(n, len(self.df), self.df['b28013'])

icf_eng = ICF_Est_New().df


def get_icf_group(code):
    # Tagastab valitud koodi koodigrupi vahemiku d4105 -> d410-d429
    codestr = code[:4]
    while int(icf_eng[codestr]['levelno']) > 3 and icf_eng[codestr]['parent']:
        codestr = icf_eng[codestr]['parent']
    # print(code, codestr, icf_eng[codestr]['Translated_title'])
    return codestr

def get_rfk_title(code):
    if code[0] in icf_eng.keys():
        try:
            title = icf_eng[code]['Translated_title']
            if not title and icf_eng[code]['parent']:
                title = get_rfk_title(icf_eng[code]['parent'])
        except KeyError: # ilmselt osaline kood nt b23
            title = ''
    else:
        title = 'Täpsustamata'
    return title

def get_rfk_qualifier(code, score, score_level=1):
    qualifiers = [
        { # 1. määraja
            'y': ['ei ole probleemi', 'kerge probleem', 'mõõdukas probleem', 'raske probleem', 'täielik probleem'],
            'd': ['ei ole piirangut', 'kerge piirang', 'mõõdukas piirang', 'raske piirang', 'täielik piirang'],
            'b': ['ei ole häiret', 'kerge häire', 'mõõdukas häire', 'raske häire', 'täielik häire'],
            's': ['ei ole kahjustust', 'kerge kahjustus', 'mõõdukas kahjustus', 'raske kahjustus', 'täielik kahjustus'],
        },
        { # 2. määraja
            'd': ['ei ole piirangut', 'kerge piirang', 'mõõdukas piirang', 'raske piirang', 'täielik piirang'],
            's': ['0', '1', '2', '3', '4'],
        },
        { # 3. määraja
            's': ['ei ole kahjustust', 'kerge kahjustus', 'mõõdukas kahjustus', 'raske kahjustus', 'täielik kahjustus'],
        }
    ]
    try:
        qualifier = qualifiers[score_level-1][code[0][0]][score]
    except IndexError:
        qualifier = score
    return qualifier

def highlight_matches(phrases, string):
    pattern = '|'.join(phrases)
    junks = re.finditer(pattern, string.lower())
    string_formatted = ''
    lastposition = 0
    while True:
        try:
            junk = next(junks)
            string_formatted += string[lastposition:junk.start()]
            string_formatted += f'<span class="highlight">{string[junk.start():junk.end()]}</span>'
            lastposition = junk.end()
        except StopIteration:
            string_formatted += string[lastposition:]
            break
    return string_formatted

# Tagastab otsingufraaside alusel kõik sobivad RFK koodid
def get_icf_matches(request=None, q=''):
    # Otsib v2ljadelt:
    # "code": "b2301"
    # "Translated_title": "Helide eristamine",
    # "Translated_description": pikk kirjeldus
    if request:
        q = request.GET.get('q', '')
    phrases = q.lower().split(' ')
    matches = []
    for key in icf_eng.keys():
        slug = ' '.join([icf_eng[key][field].lower() for field in ['code', 'Translated_title', 'Translated_description']])
        if all([(slug.find(phrase) > -1) for phrase in phrases]):
            code = icf_eng[key]['code']
            code_translated = highlight_matches(phrases, code)
            translated_title = icf_eng[key]['Translated_title']
            translated_title_formatted = highlight_matches(phrases, translated_title)
            translated_description = icf_eng[key]['Translated_description']
            translated_description_formatted = highlight_matches(phrases, translated_description)
            translated_inclusions = icf_eng[key]['Translated_inclusions']
            translated_exclusions = icf_eng[key]['Translated_exclusions']
            matches.append(
                (
                    f"""
                    <a
                              class="copylink"
                              data-bs-toggle="tooltip" data-bs-placement="top" title="kopeeri '{code} {translated_title}' lõikelauale"
                              onclick="getLinkCopy(this)"
                              id="{code}"
                              data-uri="{code} {translated_title}"
                    >
                        <strong>{code_translated}</strong>
                        <span
                            class="copylink-confirmation"
                            id="linkCopyTooltip_{code}"
                        >
                            <svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" fill="currentColor" class="bi bi-clipboard" viewBox="0 0 16 16">
                                <path d="M4 1.5H3a2 2 0 0 0-2 2V14a2 2 0 0 0 2 2h10a2 2 0 0 0 2-2V3.5a2 2 0 0 0-2-2h-1v1h1a1 1 0 0 1 1 1V14a1 1 0 0 1-1 1H3a1 1 0 0 1-1-1V3.5a1 1 0 0 1 1-1h1z"/>
                                <path d="M9.5 1a.5.5 0 0 1 .5.5v1a.5.5 0 0 1-.5.5h-3a.5.5 0 0 1-.5-.5v-1a.5.5 0 0 1 .5-.5zm-3-1A1.5 1.5 0 0 0 5 1.5v1A1.5 1.5 0 0 0 6.5 4h3A1.5 1.5 0 0 0 11 2.5v-1A1.5 1.5 0 0 0 9.5 0z"/>
                            </svg>
                        </span>
                    </a>
                    """,
                    translated_title_formatted,
                    translated_description_formatted,
                    translated_inclusions,
                    translated_exclusions,
                    # ' '.join([code, translated_title])
                )
            )
    if request:
        return JsonResponse(
            {
                'result': {
                    'items': len(matches),
                    'matches': matches[:100]
                },
            },
            safe=False
        )
    else:
        return matches

# Kirjutab lahti RFK koodi määrajad
def get_icf_qualifier(group, category, direction, scores):
    qualifiers = [
        { # 1. määraja
            'd.': {
                '_': '',
                '0': 'ei ole piirangut soorituses',
                '1': 'kerge piirang soorituses',
                '2': 'mõõdukas piirang soorituses',
                '3': 'raske piirang soorituses',
                '4': 'täielik piirang soorituses',
                '8': 'täpsustamata sooritus',
                '9': 'ei ole rakendatav'
            },
            'b.': {
                '_': '',
                '0': 'ei ole häiret',
                '1': 'kerge häire',
                '2': 'mõõdukas häire',
                '3': 'raske häire',
                '4': 'täielik häire',
                '8': 'täpsustamata',
                '9': 'ei ole rakendatav'
            },
            's.': {
                '_': '',
                '0': 'ei ole kahjustust',
                '1': 'kerge kahjustus',
                '2': 'mõõdukas kahjustus',
                '3': 'raske kahjustus',
                '4': 'täielik kahjustus',
                '8': 'täpsustamata',
                '9': 'ei ole rakendatav'
            },
            'e.': {
                '_': '',
                '0': 'ei ole takistavat tegurit',
                '1': 'kerge takistav tegur',
                '2': 'mõõdukas takistav tegur',
                '3': 'raske takistav tegur',
                '4': 'täielik takistav tegur',
                '8': 'täpsustamata',
                '9': 'ei ole rakendatav'
            },
            'e+': {
                '_': '',
                '0': 'ei ole hõlbustavat tegurit',
                '1': 'kerge hõlbustav tegur',
                '2': 'mõõdukas hõlbustav tegur',
                '3': 'raske hõlbustav tegur',
                '4': 'täielik hõlbustav tegur',
                '8': 'täpsustamata',
                '9': 'ei ole rakendatav'
            },
        },
        { # 2. määraja
            'd.': {
                '_': '',
                '0': 'ei ole piirangut suutlikkuses',
                '1': 'kerge piirang suutlikkuses',
                '2': 'mõõdukas piirang suutlikkuses',
                '3': 'raske piirang suutlikkuses',
                '4': 'täielik piirang suutlikkuses',
                '8': 'täpsustamata suutlikkus',
                '9': 'ei ole rakendatav'
            },
            's.': {
                '_': '',
                '0': 'ei ole struktuurimuutust',
                '1': 'struktuuri täielik puudumine',
                '2': 'osaline puudumine',
                '3': 'lisaosa',
                '4': 'hälbinud mõõtmed',
                '5': 'diskontinuiteet, ebapidevus',
                '6': 'väärasend',
                '7': 'kvalitatiivsed struktuurimuudatused',
                '8': 'täpsustamata',
                '9': 'ei ole rakendatav'
            },
        },
        { # 3. määraja
            's.': {
                '_': '',
                '0': 'rohkem kui üks piirkond',
                '1': 'paremal',
                '2': 'vasakul',
                '3': 'mõlemapoolne',
                '4': 'ees',
                '5': 'taga',
                '6': 'proksimaalne (keha keskteljele lähem)',
                '7': 'distaalne (keha keskteljele kaugeem)',
                '8': 'täpsustamata',
                '9': 'ei ole rakendatav'
            },
        }
    ]
    result = []
    for n in range(len(scores)):
        try:
            score = qualifiers[n][group+direction][scores[n]]
        except:
            score = ''
        if score:
            result.append(f'<span class="icf-qualifier-{n+1}"> {score}</span>')
        elif scores[n] == '_':
            pass
        else:
            result = []
            break
    return ' '.join(result)

# get_icf_code_formatted(*result.groups())
def get_icf_code_colored(group, category, direction, scores, flat=False):
    if flat:
        icf_code_colored = ''.join(
            [
                f'{group}{category}',
                f'{direction}{scores[0]}',
                f'{scores[1]}' if len(scores) > 1 else '',
                f'{scores[2]}' if len(scores) > 2 else '',
            ]
        )
    else:
        icf_code_colored = ''.join(
            [
                f'<span class="icf-category">{group}{category}</span>',
                f'<span class="icf-qualifier-1">{direction}{scores[0]}</span>',
                f'<span class="icf-qualifier-2">{scores[1]}</span>' if len(scores) > 1 else '',
                f'<span class="icf-qualifier-3">{scores[2]}</span>' if len(scores) > 2 else '',
            ]
        )
    return icf_code_colored

# Kirjutab lahti RFK koodi
def get_icf_code_verbose(request=None, code='', flat=False):
    """
    code RFK kood koos möärajaga
    flat=False HTML tage ei kasutata
    """
    if request:
        code = request.GET.get('code', '')
        code = code.strip().lower()
    pattern = r'^([b|d|e|s])([0-9]{3,5})(\.|\+)([_|0-9]{1,3})'
    result = re.search(pattern, code)
    icf_code_verbose = ''
    if result:
        category = ''.join(result.groups()[:2])
        category_verbose = get_rfk_title(category)
        qualifiers_verbose = get_icf_qualifier(*result.groups())
        if category_verbose and qualifiers_verbose:
            code_formatted = get_icf_code_colored(*result.groups(), flat=flat)
            if flat:
                icf_code_verbose = f'<span class="icf-code">{code_formatted}</span>: {category_verbose} - {qualifiers_verbose}'
            else:
                icf_code_verbose = f'<span class="icf-code" style="color: green;">{code_formatted}</span>: {category_verbose} - {qualifiers_verbose}'
    if request:
        return JsonResponse(
            {
                'result': icf_code_verbose
            },
            safe=False
        )
    else:
        return icf_code_verbose

def get_icf_path(request=None, code=None):
    if request:
        code = request.GET.get('code', '')
    codestr = f'{code} {get_rfk_title(code)}'
    if code[0] in icf_eng.keys():
        try:
            parent = icf_eng[code]['parent']
            while parent:
                codestr = '->'.join([codestr, f'{parent} {get_rfk_title(parent)}'])
                parent = icf_eng[parent]['parent']
        except KeyError: # ilmselt osaline kood nt b23
            codestr = ''
    else:
        codestr = 'Täpsustamata'
    if request:
        return JsonResponse(
            {
                'rfkPath': codestr,
            },
            safe=False
        )
    else:
        return codestr

#
# Vaated
#
def index(request):
    return render(
        request,
        'main/index.html',
        {}
    )

def privacy(request):
    return render(
        request,
        'main/privacy.html',
        {}
    )

def covidpass_s9a(request):
    print(request.COOKIES)
    html = render(
        request,
        'main/covidpass_s9a.html',
        {
            # 'object': 'Seda saab jagada!'
        }
    )
    return HttpResponse(html)

import qrcode
import qrcode.image.svg
import io
from io import BytesIO
from qrcode.image.styledpil import StyledPilImage
import base64
# import modules
from PIL import Image, ImageOps

def covidpass_s9a_3(request):
    context = {}
    if request.method == "POST":
        factory = qrcode.image.svg.SvgImage
        img = qrcode.make(
            request.POST.get("qr_text", ""),
            image_factory=factory,
            box_size=20,
            # embeded_image_path = "main/covidpass_S9a.png"
        )
        stream = BytesIO()
        img.save(stream)
        context["svg"] = stream.getvalue().decode()

        qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_L)
        qr.add_data(request.POST.get("qr_text", ""))

        # img_1 = qr.make_image(image_factory=StyledPilImage, module_drawer=RoundedModuleDrawer())
        # img_2 = qr.make_image(image_factory=StyledPilImage, color_mask=RadialGradiantColorMask())

    return render(request, "main/covidpass_s9a_2.html", context=context)

def covidpass_s9a_2(request):
    # taking image which user wants
    # in the QR code center
    Logo_link = 'main/static/main/android-chrome-192x192.png'

    logo = Image.open(Logo_link)
    new_image = Image.new("RGBA", logo.size, "WHITE")  # Create a white rgba background
    new_image.paste(logo, (0, 0), logo)  # Paste the image on the background. Go to the links given below for details.
    # new_image.convert('RGB').save('test.jpg', "JPEG")  # Save as JPEG
    logo = new_image

    # taking base width
    basewidth = 100

    # adjust image size
    wpercent = (basewidth / float(logo.size[0]))
    hsize = int((float(logo.size[1]) * float(wpercent)))
    logo = logo.resize((basewidth-30, hsize-30), Image.ANTIALIAS)
    logo = ImageOps.expand(logo, border=10, fill='white')
    QRcode = qrcode.QRCode(
        error_correction=qrcode.constants.ERROR_CORRECT_H
    )

    # taking url or text
    url = 'https://valgalinn.ee/wiki/isik/62-johan-m%C3%BCllerson/'

    # adding URL or text to QRcode
    QRcode.add_data(url)

    # generating QR code
    QRcode.make()

    # taking color name from user
    QRcolor = 'Green'

    # adding color to QR code
    QRimg = QRcode.make_image(
        fill_color=QRcolor, back_color="white").convert('RGB')

    # set size of QR code
    pos = ((QRimg.size[0] - logo.size[0]) // 2,
           (QRimg.size[1] - logo.size[1]) // 2)
    QRimg.paste(logo, pos)

    # save the QR code generated
    # QRimg.save('gfg_QR.png')

    # print('QR code generated!')
    stream = BytesIO()
    QRimg.save(stream, "PNG")

    image_data = base64.b64encode(stream.getvalue()).decode('utf-8')

    # return HttpResponse(f'<img id="plt" src="data:image/png;base64, {image_data}"></img>')
    context = {'img': f'<img id="plt" src="data:image/png;base64, {image_data}"></img>'}
    return render(request, "main/covidpass_s9a_2.html", context=context)



def sandbox(request):
    return render(
        request,
        'main/sandbox.html',
        {}
    )

def rfk(request):
    return render(
        request,
        'main/rfk.html',
        {}
    )

# Loeb textareast kõik RFK koodid koos määrajatega listi ja eraldab komponentideks
def read_content_to_rfk(icf_eng_set, content, method=1):
    data = re.findall(RFK_REGEX, content)
    if method == 4: # 0-määrajatega koode ei arvestata
        start = 1
    else:
        start = 0
    codeset = dict()
    for el in data:
        code = el.strip().split('.')
        if code[0] in icf_eng_set.keys():
            if code[1][0] in [str(m22raja) for m22raja in range(start, 5)]:
                codeset[code[0]] = (
                    code[0][:2], # 1. taseme kood d4
                    code[0][:4], # 2. taseme kood d410
                    get_icf_group(code[0][:4]), # koodigrupp nt d410-d429
                    int(code[1][0]) # ainult esimene määraja
                    )
            else:
                print('Viga: vale määraja: ', code)
        else:
            print('Viga: pole koodi: ', code)
    return codeset

def make_icf_table(rfk_set):
    trs = []
    for el in rfk_set:
        trs.append(tr(td(f'{el}:{rfk_set[el]}')))
    icf_table_html = table(trs)
    return icf_table_html.render()

def calc_mean(rfk_set, parts, part, code, method=1):
    if method == 1:  # aritmeetiline keskmine
        try:
            parts[part] = [
                parts[part][0] + rfk_set[code][3],
                parts[part][1] + 1,
                ' + '.join([f'{code}.{rfk_set[code][3]}', parts[part][2]])
            ]
        except:
            parts[part] = [rfk_set[code][3], 1, f'{code}.{rfk_set[code][3]}']
    elif method == 2:  # ruutkeskmine
        try:
            parts[part] = [
                parts[part][0] + (rfk_set[code][3] ** 2),
                parts[part][1] + 1,
                ' + '.join([f'{code}.{rfk_set[code][3]} ^ 2', parts[part][2]])
            ]
        except:
            parts[part] = [rfk_set[code][3] ** 2, 1, f'{code}.{rfk_set[code][3]} ^ 2']
    elif method == 3:  # geomeetriline keskmine NB! 0 -> 1
        try:
            parts[part] = [
                parts[part][0] * (rfk_set[code][3] if rfk_set[code][3]>0 else 1),
                parts[part][1] + 1,
                ' * '.join([f'{code}.{rfk_set[code][3]}', parts[part][2]])
            ]
        except:
            parts[part] = [rfk_set[code][3] if rfk_set[code][3]>0 else 1, 1, f'{code}.{rfk_set[code][3]}']
    elif method == 4:  # geomeetriline keskmine NB! ignoreeritakse 0-väärtusi
        if rfk_set[code][3] > 0:
            try:
                parts[part] = [
                    parts[part][0] * rfk_set[code][3],
                    parts[part][1] + 1,
                    ' * '.join([f'{code}.{rfk_set[code][3]}', parts[part][2]])
                ]
            except:
                parts[part] = [rfk_set[code][3], 1, f'{code}.{rfk_set[code][3]}']
    return parts

def calc_score(row_code, col_code, method=1):
    # print(row_code, col_code)
    if method == 1: # aritmeetiline keskmine
        row_score = row_code[0] / row_code[1] if row_code else 0
        col_score = col_code[0] / col_code[1] if col_code else 0
        mean_count = 2 if all([row_code, col_code]) else 1
        score = round(
            (row_score + col_score) / mean_count,
            1
        )
        row_string = f'{row_code[2]}) / {row_code[1]}' if row_code else '0'
        col_string = f'{col_code[2]}) / {col_code[1]}' if col_code else '0'
        title = f'(({row_string}) + ({col_string})) / {mean_count} = {score}'
    elif method == 2: # ruutkeskmine
        row_score = math.sqrt(row_code[0] / row_code[1]) if row_code else 0
        col_score = math.sqrt(col_code[0] / col_code[1]) if col_code else 0
        mean_count = 2 if all([row_code, col_code]) else 1
        score = row_score**2 + col_score**2
        score = round(
            math.sqrt(score/mean_count),
            1
        )
        title = f'sqrt((({row_code[2]})/{row_code[1]}) + ({col_code[2]})/{col_code[1]})) / {mean_count}) = {score}'
    elif method == 3: # geomeetriline keskmine NB! 0 -> 1
        row_score = row_code[0] ** (1/row_code[1]) if row_code else 1
        col_score = col_code[0] ** (1/col_code[1]) if col_code else 1
        mean_count = 2 # if all([row_code, col_code]) else 1
        score = round(
            (row_score * col_score) ** (1/mean_count),
            1
        )
        row_string = f'{row_code[2]}) ** 1/{row_code[1]}' if row_code else '1'
        col_string = f'{col_code[2]}) ** 1/{col_code[1]}' if col_code else '1'
        title = f'(({row_string} * ({col_string}) ** 1/{mean_count} = {score}'
    elif method == 4:  # geomeetriline keskmine NB! ignoreeritakse 0-väärtusi
        row_score = row_code[0] ** (1 / row_code[1]) if row_code else 1
        col_score = col_code[0] ** (1 / col_code[1]) if col_code else 1
        mean_count = 2  # if all([row_code, col_code]) else 1
        score = round(
            (row_score * col_score) ** (1 / mean_count),
            1
        )
        row_string = f'{row_code[2]}) ** 1/{row_code[1]}' if row_code else '1'
        col_string = f'{col_code[2]}) ** 1/{col_code[1]}' if col_code else '1'
        title = f'(({row_string} * ({col_string}) ** 1/{mean_count} = {score}'
    score = int(score)
    return score, title

def make_icf_matrix(rfk_set, rows=ROWS, columns=COLUMNS, ignore=IGNORE, level=1, method=1):
    #
    # level=1 kahekohaline nt b2
    # level=2 neljakohaline nt b230
    # level=3 koodigrupp nt b230-b239
    #
    # method=1 aritmeetiline keskmine
    # method=2 geomeetriline keskmine
    # method=3 ruutkeskmine
    #
    parts = dict()
    for code in rfk_set:
        if level == 0 or level == 1:
            part = code[:2] # kahekohaline nt b2
        elif level == 2:
            part = code[:4] # neljakohaline nt b230
        else: # level == 3
            part = rfk_set[code][2] # koodigrupp nt b230-b239
        if code[0] not in ignore:
            parts = calc_mean(rfk_set, parts, part, code, method)

    if level == 0: # võetakse kõik kahekohalised klassifikaatorikoodid (b1, b2, ..., d1, d2 jne)
        vect_rows = [
            code
            for code
            in icf_eng.keys()
            if (code[0] in rows and len(code) == 2)
        ]
        columns_exist = any((code[0] in columns) for code in rfk_set) # kas on func või strukt koode?
        # columns_exist = True
        if columns_exist:
            vect_columns = [
                code
                for code
                in icf_eng.keys()
                if (code[0] in columns and len(code) == 2)
            ]
        else:
            vect_columns = ['TTa']  # kui func või struct piiranguid pole, siis Täpsustamata (TTa)
    else:
        vect_rows = [
            code_block
            for code_block
            in parts.keys()
            if code_block[0] in rows
            ]
        if vect_rows:
            vect_rows.sort()
        else:
            vect_rows = ['TTa']  # kui tegevus/osalus piiranguid pole, siis Täpsustamata (TTa)

        vect_columns = [
            code_block
            for code_block
            in parts.keys()
            if code_block[0] in columns
            ]
        if vect_columns:
            vect_columns.sort()
        else:
            vect_columns = ['TTa'] # kui func või struct piiranguid pole, siis Täpsustamata (TTa)

    if max(map(len, vect_columns)) > 4:
        column_header_class = 'verticalColumnHeader' # kui veerupealkirjad pikad, siis vertikaalselt
    else:
        column_header_class = 'w3-center'

    header = tr(
        th(''),
        [
            th(
                div(
                    class_=column_header_class,
                    title=get_icf_path(request=None, code=col)
                )(col)
            )
            for col
            in vect_columns
        ]
    )
    trs = []
    for r in vect_rows:
        row = [td(title=get_icf_path(request=None, code=r))(r)]
        for c in vect_columns:
            title = ''
            try:
                score, title = calc_score(parts[r], parts[c], method=method)
            except:
                if c == 'TTa': # kui func/struct t2psustamata, siis ainult d keskmine
                    try:
                        score, title = calc_score(parts[r], None, method=method)
                    except KeyError:
                        score = ''
                elif r == 'TTa': # kui tegevus/osalus t2psustamata, siis ainult b/s keskmine
                    score, title = calc_score(None, parts[c], method=method)
                else:
                    score = ''
            if score:
                score_class = SCORE_CLASSES[score]
            else:
                score_class = ''
            el = str(score)

            row.append(td(class_=f"w3-center {score_class}", title=title)(el))
        trs.append(tr(row))
    return table(border="1", class_="w3-table-all w3-small w3-card-4")(header, trs).render()

def make_icf_verbose(rfk_set, rows=ROWS, columns=COLUMNS_VERBOSE, ignore=IGNORE_VERBOSE, level=1, method=1):
    #
    # level=1 kahekohaline nt b2
    # level=2 neljakohaline nt b230
    # level=3 koodigrupp nt b230-b239
    #
    # method=1 aritmeetiline keskmine
    # method=2 geomeetriline keskmine
    # method=3 ruutkeskmine
    #
    parts = dict()
    for code in rfk_set:
        if level == 0 or level == 1:
            part = code[:2] # kahekohaline nt b2
        elif level == 2:
            part = code[:4] # neljakohaline nt b230
        elif level == 3:
            part = rfk_set[code][2] # koodigrupp nt b230-b239
        else:
            part = code # kood t2ielikult nt b2301
        if code[0] not in ignore:
            parts = calc_mean(rfk_set, parts, part, code, method)

    if level == 0: # võetakse kõik kahekohalised klassifikaatorikoodid (b1, b2, ..., d1, d2 jne)
        vect_rows = [
            code
            for code
            in icf_eng.keys()
            if (code[0] in rows and len(code) == 2)
        ]
        columns_exist = any((code[0] in columns) for code in rfk_set) # kas on func või strukt koode?
        if columns_exist:
            vect_columns = [
                code
                for code
                in icf_eng.keys()
                if (code[0] in columns and len(code) == 2)
            ]
        else:
            vect_columns = ['TTa']  # kui func või struct piiranguid pole, siis Täpsustamata (TTa)
    else:
        vect_rows = [
            code_block
            for code_block
            in parts.keys()
            if code_block[0] in rows
            ]
        if vect_rows:
            vect_rows.sort()
        else:
            vect_rows = ['TTa']  # kui tegevus/osalus piiranguid pole, siis Täpsustamata (TTa)

        vect_columns = [
            code_block
            for code_block
            in parts.keys()
            if code_block[0] in columns
            ]
        if vect_columns:
            vect_columns.sort()
        else:
            vect_columns = ['TTa'] # kui func või struct piiranguid pole, siis Täpsustamata (TTa)

    header = tr(th('kategooria'), th('määraja'))
    trs = []

    for r in [*vect_rows, *vect_columns]:
        if r == 'TTa':
            continue
        else:
            if method == 1:
                score = parts[r][0] / parts[r][1]
            elif method == 2:
                score = math.sqrt(parts[r][0] / parts[r][1])
            elif method == 3:
                score = parts[r][0] ** (1 / parts[r][1]) if parts[r][0] else 0
            elif method == 4:
                score = parts[r][0] ** (1 / parts[r][1])
            else:
                score = 9
            score = int(round(score, 1))
            title = get_rfk_title(r)
            qualifier = get_rfk_qualifier(code=r, score=score)
            try:
                score_class = SCORE_CLASSES[score]
            except IndexError:
                score_class = ''
            # print(title, qualifier, f'({r}.{score})') # row = [td(get_rfk_title(r))]
            row = tr([
                td(class_=f"{score_class}", title=title)(f'{title[:30]} ({r})'),
                td(class_=f"{score_class}", title=title)(f'{qualifier} ({score})')
            ])
            trs.append(row)
    return table(border="1", class_="w3-table-all w3-small w3-card-4")(header, trs).render()

# Vue küsib siit andmeid valdkondade jaoks
def get_icf_calcs(request):
    content = request.GET.get('content', '')
    method = int(request.GET.get('method', 1))
    rfk_set = read_content_to_rfk(icf_eng, content, method)
    # icf_table_html = make_icf_table(rfk_set)
    icf_table_matrix_level1 = make_icf_matrix(rfk_set, level=1, method=method)
    icf_table_matrix_level2 = make_icf_matrix(rfk_set, level=2, method=method)
    icf_table_matrix_level3 = make_icf_matrix(rfk_set, level=3, method=method)

    icf_table_verbose_level4 = make_icf_verbose(rfk_set, level=4, method=method)

    return JsonResponse(
        {
            # 'icf_table_html': icf_table_html,
            'icf_table_matrix_level1': icf_table_matrix_level1,
            'icf_table_matrix_level2': icf_table_matrix_level2,
            'icf_table_matrix_level3': icf_table_matrix_level3,
            'icf_table_verbose_level4': icf_table_verbose_level4,
            'rfk_codeset_count': len(rfk_set.keys())
        },
        safe=False
    )

# Vue küsib siit andmeid kokkuvõtte jaoks
def get_icf_summary(request):
    content = request.GET.get('content', '')
    method = int(request.GET.get('method', 1))
    rfk_set = read_content_to_rfk(icf_eng, content, method)
    icf_table_html = make_icf_table(rfk_set)
    icf_table_matrix_level0 = make_icf_matrix(rfk_set, level=0, method=method)
    icf_table_matrix_level1 = make_icf_matrix(rfk_set, level=1, method=method)
    icf_table_matrix_level2 = make_icf_matrix(rfk_set, level=2, method=method)
    icf_table_matrix_level3 = make_icf_matrix(rfk_set, level=3, method=method)

    icf_table_verbose_level2 = make_icf_verbose(rfk_set, level=2, method=method)
    icf_table_verbose_level3 = make_icf_verbose(rfk_set, level=3, method=method)
    icf_table_verbose_level4 = make_icf_verbose(rfk_set, level=4, method=method)

    return JsonResponse(
        {
            'icf_table_html': icf_table_html,
            'icf_table_matrix_level0': icf_table_matrix_level0,
            'icf_table_matrix_level1': icf_table_matrix_level1,
            'icf_table_matrix_level2': icf_table_matrix_level2,
            'icf_table_matrix_level3': icf_table_matrix_level3,
            'icf_table_verbose_level2': icf_table_verbose_level2,
            'icf_table_verbose_level3': icf_table_verbose_level3,
            'icf_table_verbose_level4': icf_table_verbose_level4
        },
        safe=False
    )

def is_code_in_group(code, group):
    while icf_eng[code] and int(icf_eng[code]['levelno']) > 1 and icf_eng[code]['parent']:
        if (code == group) or (code == icf_eng[code]['parent']):
            return True
        code = icf_eng[code]['parent']
    return False



def test(method=1):
    level = 1
    rfk_set_b = {
        'b4551': ('b4', 'b455', 'b450-b469', 2),
        'b4552': ('b4', 'b455', 'b450-b469', 3),
        'b4553': ('b4', 'b455', 'b450-b469', 2),
    }
    rfk_set_d = {
        'd4551': ('d4', 'd455', 'd450-b469', 0),
        'd4552': ('d4', 'd455', 'd450-b469', 3),
        'd4553': ('d4', 'd455', 'd450-b469', 2),
    }
    rfk_set = dict()
    for key in rfk_set_b:
        rfk_set[key] = rfk_set_b[key]
    for key in rfk_set_d:
        rfk_set[key] = rfk_set_d[key]
    parts = dict()
    for code in rfk_set:
        part = code[:2]  # kahekohaline nt b2
        parts = calc_mean(rfk_set, parts, part, code, method)

    rows = 'd'
    columns = 'b'

    vect_rows = [
        code_block
        for code_block
        in parts.keys()
        if code_block[0] in rows
    ]
    if vect_rows:
        vect_rows.sort()
    else:
        vect_rows = ['TTa']  # kui tegevus/osalus piiranguid pole, siis Täpsustamata (TTa)

    vect_columns = [
        code_block
        for code_block
        in parts.keys()
        if code_block[0] in columns
    ]
    if vect_columns:
        vect_columns.sort()
    else:
        vect_columns = ['TTa']  # kui func või struct piiranguid pole, siis Täpsustamata (TTa)

    for r in vect_rows:
        for c in vect_columns:
            score = ''
            title = ''
            try:
                score, title = calc_score(parts[r], parts[c], method=method)
            except:
                if c == 'TTa': # kui func/struct t2psustamata, siis ainult d keskmine
                    try:
                        score, title = calc_score(parts[r], None, method=method)
                    except KeyError:
                        pass
                elif r == 'TTa': # kui tegevus/osalus t2psustamata, siis ainult b/s keskmine
                    score, title = calc_score(None, parts[c], method=method)
            print(score, title)

def get_client_ip(request):
    remote_address = request.META.get('HTTP_X_FORWARDED_FOR') or request.META.get('REMOTE_ADDR')
    ip = request.META.get('REMOTE_ADDR')
    # x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    # if x_forwarded_for:
    #     proxies = x_forwarded_for.split(',')
    #     while (len(proxies) > 0 and proxies[0].startswith(PRIVATE_IPS_PREFIX)):
    #         proxies.pop(0)
    #         if len(proxies) > 0:
    #             ip = proxies[0]
    return ip

def get_kysimustik7():
    filename = 'kysimustik_v3.xml'
    file = STATIC_DIR / 'data' / filename
    with open(file, 'r', encoding='utf8') as f:
        tree = ET.parse(f)
        root = tree.getroot()

    vanusgrupid = []
    vanusgruppideMuutumatudSeisundid = {}
    vanusgruppideKysimused = {}
    vanusgruppideYldKysimused = {}
    vanusgruppideFailiTekstid = {}

    n = 0
    for child in root:
        print(child.tag, child.attrib, child.find('muutumatudSeisundid').text.split('\n'))
        # vanusgruppide nimetused
        vanusgrupp = {
            'kysimustik': True if child.attrib['kysimustik'] == 'true' else False,
            'text': child.attrib['name'],
            'value': n
        }
        vanusgrupid.append(vanusgrupp)

        # vanusgruppide muutumatute seisundite loetelud
        i = itertools.count(0)
        muutumatudSeisundidData = child.find('muutumatudSeisundid')
        muutumatudSeisundidQuestion = muutumatudSeisundidData.attrib['question']
        muutumatudSeisundidList = [
            {'text': seisund.strip(), 'id': next(i)}
            for seisund
            in muutumatudSeisundidData.text.split('\n')
            if len(seisund.strip()) > 0
        ]
        vanusgruppideMuutumatudSeisundid[n] = {
            'muutumatudSeisundidQuestion': muutumatudSeisundidQuestion,
            'muutumatudSeisundidList': muutumatudSeisundidList
        }

        # vanusgrupi kysimused
        if child.attrib['kysimustik'] == 'true':
            vanusgrupiKysimusedData = child.find('vanusgrupiKysimused')
            vanusgrupiKysimusedQuestion = vanusgrupiKysimusedData.attrib['question']
            vanusgrupiKysimusedList = [
                {'text': kysimus.text.strip(), 'valdkond': kysimus.attrib['valdkond_nr'], 'score': '', 'answer': '' }
                for kysimus
                in vanusgrupiKysimusedData.findall('kysimus')
                if len(kysimus.text.strip()) > 0
            ]
        else:
            vanusgrupiKysimusedQuestion = ''
            vanusgrupiKysimusedList = []

        vanusgruppideKysimused[n] = {
            'vanusgrupiKysimusedQuestion': vanusgrupiKysimusedQuestion,
            'vanusgrupiKysimusedList': vanusgrupiKysimusedList
        }

        # vanusgrupi yldkysimused
        vanusgrupiYldKysimusedData = child.find('vanusgrupiYldKysimused')
        vanusgrupiYldKysimusedQuestion = vanusgrupiYldKysimusedData.attrib['question']
        vanusgrupiYldKysimusedList = [
            {'text': kysimus.strip(), 'answer': ''}
            for kysimus
            in vanusgrupiYldKysimusedData.text.split('\n')
            if len(kysimus.strip()) > 0
        ]
        vanusgruppideYldKysimused[n] = {
            'vanusgrupiYldKysimusedQuestion': vanusgrupiYldKysimusedQuestion,
            'vanusgrupiYldKysimusedList': vanusgrupiYldKysimusedList
        }
        vanusgrupiFailiTekst = child.find('vanusgrupiFailiTekst')
        vanusgruppideFailiTekstid[n] = vanusgrupiFailiTekst.text.strip()
        n += 1

    kysimustik = {
        'vanusgrupid': vanusgrupid,
        'vanusgruppideMuutumatudSeisundid': vanusgruppideMuutumatudSeisundid,
        'vanusgruppideKysimused': vanusgruppideKysimused,
        'vanusgruppideYldKysimused': vanusgruppideYldKysimused,
        'vanusgruppideFailiTekstid': vanusgruppideFailiTekstid
    }
    return kysimustik

# Vue kysimustik7 tulemuste salvestamiseks
def save_kysimustik7_results(request):
    kysimustik_results = {
        'version': '7',
        'ipAddress': request.GET.get('ipAddress', ''),
        'vanusgrupp': json.loads(request.GET.get('vanusgrupp', '')),
        'checkedMuutumatudSeisundid': request.GET.get('checkedMuutumatudSeisundid', ''),
        'toggleShowForm': request.GET.get('toggleShowForm', ''),
        'kysimustikList': json.loads(request.GET.get('kysimustikList', '')),
        'yldkysimusedList': json.loads(request.GET.get('yldkysimusedList', '')),
        'feedback': request.GET.get('feedback', '')
    }
    # print(kysimustik_results)
    versioon = kysimustik_results['version']
    vanusgrupp = kysimustik_results['vanusgrupp']['text']
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'eneseHinnang_v{versioon}_{vanusgrupp}_{timestamp}.json'
    with open(STATIC_DIR / 'data' / filename, 'w', encoding='utf8') as f:
        json.dump(kysimustik_results, f)

    return JsonResponse(
        {'filename': filename},
        safe=False
    )

#
# Küsimustiku vaade ver 7
#
def kysimustik7(request):
    kysimustik = get_kysimustik7()

    vanusgrupid = kysimustik['vanusgrupid']
    vanusgruppideMuutumatudSeisundid = kysimustik['vanusgruppideMuutumatudSeisundid']
    vanusgruppideKysimused = kysimustik['vanusgruppideKysimused']
    vanusgruppideYldKysimused = kysimustik['vanusgruppideYldKysimused']
    vanusgruppideFailiTekstid = kysimustik['vanusgruppideFailiTekstid']

    context = {
        'selectedSkaala': 1,
        'vanusgrupid': json.dumps(vanusgrupid),
        'vanusgruppideMuutumatudSeisundid': json.dumps(vanusgruppideMuutumatudSeisundid),
        'vanusgruppideKysimused': json.dumps(vanusgruppideKysimused),
        'vanusgruppideYldKysimused': json.dumps(vanusgruppideYldKysimused),
        'vanusgruppideFailiTekstid': json.dumps(vanusgruppideFailiTekstid),
        'ip': get_client_ip(request)
    }
    return render(
        request,
        'main/kysimustik7.html',
        context
    )

def get_kysimustik8():
    filename = 'kysimustik_v3.xml'
    file = STATIC_DIR / 'data' / filename
    with open(file, 'r', encoding='utf8') as f:
        tree = ET.parse(f)
        root = tree.getroot()

    vanusgrupid = []
    vanusgruppideMuutumatudSeisundid = {}
    vanusgruppideV6tmetegevused = {}
    vanusgruppideKysimused = {}
    vanusgruppideYldKysimused = {}
    vanusgruppideFailiTekstid = {}

    n = 0
    for child in root:
        print(child.tag, child.attrib, child.find('muutumatudSeisundid').text.split('\n'))
        # vanusgruppide nimetused
        vanusgrupp = {
            'kysimustik': True if child.attrib['kysimustik'] == 'true' else False,
            'text': child.attrib['name'],
            'value': n
        }
        vanusgrupid.append(vanusgrupp)

        # vanusgruppide muutumatute seisundite loetelud
        i = itertools.count(0)
        muutumatudSeisundidData = child.find('muutumatudSeisundid')
        muutumatudSeisundidQuestion = muutumatudSeisundidData.attrib['question']
        muutumatudSeisundidList = [
            {'text': seisund.strip(), 'id': next(i)}
            for seisund
            in muutumatudSeisundidData.text.split('\n')
            if len(seisund.strip()) > 0
        ]
        vanusgruppideMuutumatudSeisundid[n] = {
            'muutumatudSeisundidQuestion': muutumatudSeisundidQuestion,
            'muutumatudSeisundidList': muutumatudSeisundidList
        }

        # vanusgrupi v6tmetegevused
        vanusgrupiV6tmetegevusedData = child.find('vanusgrupiV6tmetegevused')
        vanusgrupiV6tmetegevusedList = [
            {
                'text': kysimus.text.strip(),
                'valdkond_nr': kysimus.attrib['valdkond_nr'],
                'v6tmetegevus_nr': kysimus.attrib['v6tmetegevus_nr'],
            }
            for kysimus
            in vanusgrupiV6tmetegevusedData.findall('v6tmetegevus')
            if len(kysimus.text.strip()) > 0
        ]

        # vanusgrupi kysimused
        if child.attrib['kysimustik'] == 'true':
            vanusgrupiKysimusedData = child.find('vanusgrupiKysimused')
            vanusgrupiKysimusedQuestion = vanusgrupiKysimusedData.attrib['question']
            vanusgrupiKysimusedList = [
                {
                    'text': kysimus.text.strip(),
                    'valdkond_nr': kysimus.attrib['valdkond_nr'],
                    'v6tmetegevus_nr': kysimus.attrib['v6tmetegevus_nr'],
                    'rfk_kategooria': kysimus.attrib['rfk_kategooria'],
                    'score': '',
                    'answer': ''
                }
                for kysimus
                in vanusgrupiKysimusedData.findall('kysimus')
                if len(kysimus.text.strip()) > 0
            ]
        else:
            vanusgrupiKysimusedQuestion = ''
            vanusgrupiKysimusedList = []

        vanusgruppideV6tmetegevused[n] = {
            'vanusgrupiV6tmetegevusedList': vanusgrupiV6tmetegevusedList
        }

        vanusgruppideKysimused[n] = {
            'vanusgrupiKysimusedQuestion': vanusgrupiKysimusedQuestion,
            'vanusgrupiKysimusedList': vanusgrupiKysimusedList
        }

        # vanusgrupi yldkysimused
        vanusgrupiYldKysimusedData = child.find('vanusgrupiYldKysimused')
        vanusgrupiYldKysimusedQuestion = vanusgrupiYldKysimusedData.attrib['question']
        vanusgrupiYldKysimusedList = [
            {
                'text': kysimus.strip(),
                'answer': ''
            }
            for kysimus
            in vanusgrupiYldKysimusedData.text.split('\n')
            if len(kysimus.strip()) > 0
        ]
        vanusgruppideYldKysimused[n] = {
            'vanusgrupiYldKysimusedQuestion': vanusgrupiYldKysimusedQuestion,
            'vanusgrupiYldKysimusedList': vanusgrupiYldKysimusedList
        }
        vanusgrupiFailiTekst = child.find('vanusgrupiFailiTekst')
        vanusgruppideFailiTekstid[n] = vanusgrupiFailiTekst.text.strip()
        n += 1

    kysimustik = {
        'vanusgrupid': vanusgrupid,
        'vanusgruppideMuutumatudSeisundid': vanusgruppideMuutumatudSeisundid,
        'vanusgruppideV6tmetegevused': vanusgruppideV6tmetegevused,
        'vanusgruppideKysimused': vanusgruppideKysimused,
        'vanusgruppideYldKysimused': vanusgruppideYldKysimused,
        'vanusgruppideFailiTekstid': vanusgruppideFailiTekstid
    }
    return kysimustik

#
# Küsimustiku vaade ver 8
#
def kysimustik8(request):
    kysimustik = get_kysimustik8()
    vanusgrupid = kysimustik['vanusgrupid']
    vanusgruppideMuutumatudSeisundid = kysimustik['vanusgruppideMuutumatudSeisundid']
    vanusgruppideV6tmetegevused = kysimustik['vanusgruppideV6tmetegevused']
    vanusgruppideKysimused = kysimustik['vanusgruppideKysimused']
    vanusgruppideYldKysimused = kysimustik['vanusgruppideYldKysimused']
    vanusgruppideFailiTekstid = kysimustik['vanusgruppideFailiTekstid']

    context = {
        'selectedSkaala': 1,
        'vanusgrupid': json.dumps(vanusgrupid),
        'vanusgruppideMuutumatudSeisundid': json.dumps(vanusgruppideMuutumatudSeisundid),
        'vanusgruppideV6tmetegevused': json.dumps(vanusgruppideV6tmetegevused),
        'vanusgruppideKysimused': json.dumps(vanusgruppideKysimused),
        'vanusgruppideYldKysimused': json.dumps(vanusgruppideYldKysimused),
        'vanusgruppideFailiTekstid': json.dumps(vanusgruppideFailiTekstid),
        'ip': get_client_ip(request)
    }
    return render(
        request,
        'main/kysimustik8.html',
        context
    )

# Vue kysimustik7 tulemuste salvestamiseks
def get_kysimustik8_ekspertiis(request):
    vanusgrupp = json.loads(request.GET.get('vanusgrupp', ''))
    checkedMuutumatudSeisundid = request.GET.get('checkedMuutumatudSeisundid', '')
    v6tmetegevusedList = json.loads(request.GET.get('v6tmetegevusedList', ''))
    kysimustikList = json.loads(request.GET.get('kysimustikList', ''))
    yldkysimusedList = json.loads(request.GET.get('yldkysimusedList', ''))

    v6tmetegevused = {
        (v6tmetegevus["valdkond_nr"], v6tmetegevus["v6tmetegevus_nr"]): v6tmetegevus["text"]
        for v6tmetegevus
        in v6tmetegevusedList
    }
    skoorid = [
        kysimus['score']
        for kysimus
        in kysimustikList
        if (kysimus['valdkond_nr'] == '1' and kysimus['v6tmetegevus_nr'] == '2')
    ]
    header = tr(
        th('võtmetegevused'), th('EH'), th('EA')
    )
    trs = []
    for key, item in v6tmetegevused.items():
        skoorid = [
            str(kysimus['score'])
            for kysimus
            in kysimustikList
            if (kysimus['valdkond_nr'] == str(key[0]) and kysimus['v6tmetegevus_nr'] == str(key[1]))
        ]
        if skoorid:
            max_skoor = max(skoorid)
        else:
            max_skoor = ''
        options = []
        for n in range(5):
            if max_skoor == str(n):
                options.append(option(value=f"{n}", selected="selected")(f"{n}"))
            else:
                options.append(option(value=f"{n}")(f"{n}"))
        select_field = select(name="skoor_{key[0]}_{key[1]}", id="skoor_{key[0]}_{key[1]}")(options)
        row = [td(f'{key[0]}.{key[1]} {item}'), td(max_skoor), td(select_field)]
        trs.append(tr(row))
    html = table(border="1", class_="w3-table-all w3-small w3-card-4")(header, trs).render()

    rfkd = [
        '.'.join([kysimus['rfk_kategooria'], str(kysimus['score'])])
        for kysimus
        in kysimustikList
        if (kysimus['score'] != '' and kysimus['rfk_kategooria'] != 'tyhi')
    ]

    trs = [
        tr(td(Safe(get_icf_code_verbose(request=None, code=rfk)))) for rfk in rfkd
    ]
    rfk_table = table(border="1", class_="w3-table-all w3-small w3-card-4")(trs).render()

    return JsonResponse(
        {
            'vanusgrupp': vanusgrupp,
            'checkedMuutumatudSeisundid': checkedMuutumatudSeisundid,
            'kysimustikList': kysimustikList,
            'html': html,
            'rfk_table': rfk_table,
        },
        safe=False
    )

# Vue kysimustik9 tulemuste salvestamiseks
def save_kysimustik9_results(request):
    kysimustik_results = {
        'version': '9',
        'ipAddress': request.GET.get('ipAddress', ''),
        'vanusgrupp': json.loads(request.GET.get('vanusgrupp', '')),
        'toggleShowForm': request.GET.get('toggleShowForm', ''),
        'kysimustikList': json.loads(request.GET.get('kysimustikList', '')),
        'yldkysimusedList': json.loads(request.GET.get('yldkysimusedList', '')),
        'feedback': request.GET.get('feedback', '')
    }
    # print(kysimustik_results)
    versioon = kysimustik_results['version']
    vanusgrupp = kysimustik_results['vanusgrupp']['text']
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'eneseHinnang_v{versioon}_{vanusgrupp}_{timestamp}.json'
    with open(STATIC_DIR / 'data' / filename, 'w', encoding='utf8') as f:
        json.dump(kysimustik_results, f)

    return JsonResponse(
        {'filename': filename},
        safe=False
    )

def get_kysimustik9():
    filename = 'kysimustik_v5.xml'
    file = STATIC_DIR / 'data' / filename
    with open(file, 'r', encoding='utf8') as f:
        tree = ET.parse(f)
        root = tree.getroot()

    vanusgrupid = []
    vanusgruppideMuutumatudSeisundid = {}
    vanusgruppideV6tmetegevused = {}
    vanusgruppideKysimused = {}
    vanusgruppideYldKysimused = {}
    vanusgruppideFailiTekstid = {}

    n = 0
    for child in root.find('vanusgrupid'):
        # print(child.tag, child.attrib, child.find('muutumatudSeisundid').text.split('\n'))
        # vanusgruppide nimetused
        vanusgrupp = {
            'kysimustik': child.attrib['kysimustik'] == 'true',
            'text': child.attrib['name'],
            'value': n
        }
        vanusgrupid.append(vanusgrupp)

        # vanusgruppide muutumatute seisundite loetelud
        # i = itertools.count(0)
        muutumatudSeisundidData = child.find('muutumatudSeisundid')
        muutumatudSeisundidQuestion = {'kysimus': muutumatudSeisundidData.text.strip(), 'answer': ''}
        vanusgruppideMuutumatudSeisundid[n] = {
            'muutumatudSeisundidQuestion': muutumatudSeisundidQuestion,
        }

        # vanusgrupi v6tmetegevused
        vanusgrupiV6tmetegevusedData = child.find('vanusgrupiV6tmetegevused')
        vanusgrupiV6tmetegevusedList = [
            {
                'text': kysimus.text.strip(),
                'valdkond_nr': kysimus.attrib['valdkond_nr'],
                'v6tmetegevus_nr': kysimus.attrib['v6tmetegevus_nr'],
            }
            for kysimus
            in vanusgrupiV6tmetegevusedData.findall('v6tmetegevus')
            if len(kysimus.text.strip()) > 0
        ]

        # vanusgrupi kysimused
        if child.attrib['kysimustik'] == 'true':
            vanusgrupiKysimusedData = child.find('vanusgrupiKysimused')
            vanusgrupiKysimusedQuestion = vanusgrupiKysimusedData.attrib['question']
            vanusgrupiKysimusedList = [
                {
                    'text': kysimus.text.strip(),
                    'valdkond_nr': kysimus.attrib['valdkond_nr'],
                    'v6tmetegevus_nr': kysimus.attrib['v6tmetegevus_nr'],
                    'rfk_kategooria': kysimus.attrib['rfk_kategooria'],
                    'kohustuslik': kysimus.attrib['kohustuslik'] == 'jah',
                    'score': '',
                    'answer': ''
                }
                for kysimus
                in vanusgrupiKysimusedData.findall('kysimus')
                if len(kysimus.text.strip()) > 0
            ]
        else:
            vanusgrupiKysimusedQuestion = ''
            vanusgrupiKysimusedList = []

        vanusgruppideV6tmetegevused[n] = {
            'vanusgrupiV6tmetegevusedList': vanusgrupiV6tmetegevusedList
        }

        vanusgruppideKysimused[n] = {
            'vanusgrupiKysimusedQuestion': vanusgrupiKysimusedQuestion,
            'vanusgrupiKysimusedList': vanusgrupiKysimusedList
        }

        # vanusgrupi yldkysimused
        vanusgrupiYldKysimusedData = child.find('vanusgrupiYldKysimused')
        vanusgrupiYldKysimusedQuestion = vanusgrupiYldKysimusedData.attrib['question']
        vanusgrupiYldKysimusedList = [
            {
                'text': kysimus.text.strip(),
                'kohustuslik': kysimus.attrib['kohustuslik'] == 'jah',
                'answer': ''
            }
            for kysimus
            in vanusgrupiYldKysimusedData.findall('kysimus')
            if len(kysimus.text.strip()) > 0
        ]
        vanusgruppideYldKysimused[n] = {
            'vanusgrupiYldKysimusedQuestion': vanusgrupiYldKysimusedQuestion,
            'vanusgrupiYldKysimusedList': vanusgrupiYldKysimusedList
        }
        vanusgrupiFailiTekst = child.find('vanusgrupiFailiTekst')
        vanusgruppideFailiTekstid[n] = vanusgrupiFailiTekst.text.strip()
        n += 1

    kysimustik = {
        'vanusgrupid': vanusgrupid,
        'vanusgruppideMuutumatudSeisundid': vanusgruppideMuutumatudSeisundid,
        'vanusgruppideV6tmetegevused': vanusgruppideV6tmetegevused,
        'vanusgruppideKysimused': vanusgruppideKysimused,
        'vanusgruppideYldKysimused': vanusgruppideYldKysimused,
        'vanusgruppideFailiTekstid': vanusgruppideFailiTekstid
    }
    return kysimustik

#
# Küsimustiku vaade ver 9
#
def kysimustik9(request):
    kysimustik = get_kysimustik9()
    vanusgrupid = kysimustik['vanusgrupid']
    vanusgruppideMuutumatudSeisundid = kysimustik['vanusgruppideMuutumatudSeisundid']
    vanusgruppideV6tmetegevused = kysimustik['vanusgruppideV6tmetegevused']
    vanusgruppideKysimused = kysimustik['vanusgruppideKysimused']
    vanusgruppideYldKysimused = kysimustik['vanusgruppideYldKysimused']
    vanusgruppideFailiTekstid = kysimustik['vanusgruppideFailiTekstid']

    context = {
        'selectedSkaala': 1,
        'vanusgrupid': json.dumps(vanusgrupid),
        'vanusgruppideMuutumatudSeisundid': json.dumps(vanusgruppideMuutumatudSeisundid),
        'vanusgruppideV6tmetegevused': json.dumps(vanusgruppideV6tmetegevused),
        'vanusgruppideKysimused': json.dumps(vanusgruppideKysimused),
        'vanusgruppideYldKysimused': json.dumps(vanusgruppideYldKysimused),
        'vanusgruppideFailiTekstid': json.dumps(vanusgruppideFailiTekstid),
        'ip': get_client_ip(request)
    }
    return render(
        request,
        'main/kysimustik9.html',
        context
    )

# Vue kysimustik9 tulemuste salvestamiseks
def get_kysimustik9_ekspertiis(request):
    vanusgrupp = json.loads(request.GET.get('vanusgrupp', ''))
    v6tmetegevusedList = json.loads(request.GET.get('v6tmetegevusedList', ''))
    kysimustikList = json.loads(request.GET.get('kysimustikList', ''))
    yldkysimusedList = json.loads(request.GET.get('yldkysimusedList', ''))

    v6tmetegevused = {
        (v6tmetegevus["valdkond_nr"], v6tmetegevus["v6tmetegevus_nr"]): v6tmetegevus["text"]
        for v6tmetegevus
        in v6tmetegevusedList
    }
    skoorid = [
        kysimus['score']
        for kysimus
        in kysimustikList
        if (kysimus['valdkond_nr'] == '1' and kysimus['v6tmetegevus_nr'] == '2')
    ]
    header = tr(
        th('võtmetegevused'), th('EH'), th('EA')
    )
    trs = []
    for key, item in v6tmetegevused.items():
        skoorid = [
            str(kysimus['score'])
            for kysimus
            in kysimustikList
            if (kysimus['valdkond_nr'] == str(key[0]) and kysimus['v6tmetegevus_nr'] == str(key[1]))
        ]
        if skoorid:
            max_skoor = max(skoorid)
        else:
            max_skoor = ''
        options = []
        for n in range(5):
            if max_skoor == str(n):
                options.append(option(value=f"{n}", selected="selected")(f"{n}"))
            else:
                options.append(option(value=f"{n}")(f"{n}"))
        select_field = select(name="skoor_{key[0]}_{key[1]}", id="skoor_{key[0]}_{key[1]}")(options)
        row = [td(f'{key[0]}.{key[1]} {item}'), td(max_skoor), td(select_field)]
        trs.append(tr(row))
    html = table(border="1", class_="w3-table-all w3-small w3-card-4")(header, trs).render()

    rfkd = [
        '.'.join([kysimus['rfk_kategooria'], str(kysimus['score'])])
        for kysimus
        in kysimustikList
        if (kysimus['score'] != '' and kysimus['rfk_kategooria'] != 'tyhi')
    ]

    trs = [
        tr(td(Safe(get_icf_code_verbose(request=None, code=rfk)))) for rfk in rfkd
    ]
    rfk_table = table(border="1", class_="w3-table-all w3-small w3-card-4")(trs).render()

    return JsonResponse(
        {
            'vanusgrupp': vanusgrupp,
            'kysimustikList': kysimustikList,
            'html': html,
            'rfk_table': rfk_table,
        },
        safe=False
    )

import io
from django.http import FileResponse
from reportlab.pdfgen import canvas

def some_view(request):
    # Create a file-like buffer to receive PDF data.
    buffer = io.BytesIO()

    # Create the PDF object, using the buffer as its "file."
    p = canvas.Canvas(buffer)

    # Draw things on the PDF. Here's where the PDF generation happens.
    # See the ReportLab documentation for the full list of functionality.
    p.drawString(100, 100, "Hello world.")

    # Close the PDF object cleanly, and we're done.
    p.showPage()
    p.save()

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='hello.pdf')

import openpyxl

def get_excel():
    from django.conf import settings

    path = settings.BASE_DIR
    DATA_DIR = path / 'main' / 'static' / 'main' / 'data'

    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook(
        DATA_DIR / 'Valdkondade loogika ja vaated_LAVA.xlsx',
        read_only=False
    )
    print(dataframe.sheetnames)

    # Define variable to read sheet
    # dataframe1 = dataframe.active
    sheet = dataframe['Küsimustik']
    cols = [col[0].value for col in sheet.columns]
    print(cols)

    vanusgrupp = 'VPI'
    # Iterate the loop to read the cell values
    for row in range(1, sheet.max_row):
        cols = [col[row].value for col in sheet.iter_cols(1, sheet.max_column)]
        # print(cols)
        if cols[0] == vanusgrupp:
            valdkond_nr = cols[1]
            v6tmetegevus_nr = cols[2]
            v6tmetegevus_nimetus = cols[3]
            print(f'<v6tmetegevus valdkond_nr="{valdkond_nr}" v6tmetegevus_nr="{v6tmetegevus_nr}">{v6tmetegevus_nimetus}</v6tmetegevus>')
        # for col in sheet.iter_cols(1, sheet.max_column):
        #     print(col[row].value)

import openpyxl
from main.forms import KategooriaForm, ArticleFormSet

def read_codeset_from_excel():
    wb = openpyxl.load_workbook(settings.BASE_DIR / 'main' / 'static' / 'main' / 'data' / 'RFK_koodisett.xlsx')
    ws = wb.active
    codeset = dict()
    for row in ws.iter_rows(
            min_row=2, max_row=None,
            min_col=1, max_col=7,
            values_only=True
    ):
        codeset[row[0]] = {
            'valdkond': row[0][:2], # d1, d2 etc
            'nimetus': row[1],
            'seotud': row[2],
            'koolieelik': row[3],
            'kooliealine': row[4],
            'vanaduspensioniealine': row[5],
            'tööealine': row[6],
            'form': KategooriaForm(
                auto_id=f'{row[0]}_%s',
            )
        }
    return codeset

def prt(request):
    vanusgrupid = ['kõik', 'koolieelik', 'kooliealine', 'tööealine', 'vanaduspensioniealine']
    valitud_vanusgrupp = request.GET.get('vanusgrupp')
    codeset = read_codeset_from_excel()
    if valitud_vanusgrupp is not None and valitud_vanusgrupp in vanusgrupid:
        codeset = {k:v for k,v in codeset.items() if codeset[k][valitud_vanusgrupp] is not None}
    if request.method == "POST":
        form = KategooriaForm(request.POST)
        if form.is_valid():
            kategooria = form.cleaned_data["kategooria"]
            print(kategooria)
    else:
        pass
        form = KategooriaForm()
    return render(
        request,
        'main/prt.html',
        {
            'codeset': codeset,
            'vanusgrupid': vanusgrupid,
            'valitud_vanusgrupp': valitud_vanusgrupp
        }
    )

import random
def j6ul2023(request):
    rfk_codes_b = [
        'b535.1', # Seedesüsteemiga seonduvad aistingud
        'b1265.0', # Optimism
        'b1300.0', # Vaimse energia tase
    ]
    rfk_codes_d = [
        'd550.00', # Söömine
        'd6604.0', # Teiste abistamine toitumisel
        'd760.0', # Perekondlikud suhted
        'd350.00', # Vestlemine
        'd920.0', # Puhke - ja vabaajategevustes osalemine
        'd460.0', # Erinevates kohtades liikumine
    ]
    rfk_codes_e = [
        'e1401+4', # Kultuuri -, sporditegevuse ja vaba aja veetmise abivahendid ning tehnoloogiad
        'e350+4', # Kodustatud loomad
        'e320+4', # Sõbrad
    ]
    random.shuffle(rfk_codes_b)
    random.shuffle(rfk_codes_d)
    random.shuffle(rfk_codes_e)
    return render(
        request,
        'main/j6ul2023.html',
        {
            'snowflakes': range(1, 108),
            'rfk_code_verbose_b1': get_icf_code_verbose(request=None, code=rfk_codes_b[0], flat=False),
            'rfk_code_verbose_d1': get_icf_code_verbose(request=None, code=rfk_codes_d[0], flat=False),
            'rfk_code_verbose_d2': get_icf_code_verbose(request=None, code=rfk_codes_d[1], flat=False),
            'rfk_code_verbose_e1': get_icf_code_verbose(request=None, code=rfk_codes_e[0], flat=False),
        }
    )

from main.models import RFK
def import_icf2db():
    for code in icf_eng:
        row = RFK(**icf_eng[code])
        print(row)
        row.save()

# arvutab maatriksi level1 d x b m22rajatega
def calc_dblevel1_qualifiers(codeset):
    rfk_set = dict()
    for d_code in range(1, 10):
        rfk_set[f'd{d_code}'] = dict()
        rfk_set[f'd{d_code}']['d_qualifier'] = ''
        for b_code in range(1, 9):
            rfk_set[f'd{d_code}'][f'b{b_code}'] = ''
    for code_pair in codeset:
        if code_pair[0][0] == 'd' and code_pair[1][0] == 'b': # ainult kui on d ja b koodipaar
            d_code, d_qualifier = code_pair[0].split('.')
            b_code, b_qualifier = code_pair[1].split('.')
            if (rfk_set[d_code[:2]]['d_qualifier'] < d_qualifier):
                rfk_set[d_code[:2]]['d_qualifier'] = d_qualifier
            if (rfk_set[d_code[:2]][b_code[:2]] < b_qualifier):
                rfk_set[d_code[:2]][b_code[:2]] = b_qualifier
    return rfk_set

# arvutab maatriksi level1 d x b m22rajatega
def calc_blevel2_qualifiers(codeset):
    rfk_set = dict()
    for code_pair in codeset:
        if code_pair[0][0] == 'd' and code_pair[1][0] == 'b': # ainult kui on d ja b koodipaar
            b_code, b_qualifier = code_pair[1].split('.')
            blevel2_code = get_icf_group(b_code)
            if blevel2_code in rfk_set.keys():
                if rfk_set[blevel2_code] < b_qualifier:
                    rfk_set[blevel2_code] = b_qualifier
            else:
                rfk_set[blevel2_code] = b_qualifier
    return rfk_set

def calc_dlevel1_max(db_level1_dict):
    qualifier = max(
        [
            db_level1_dict[key]["d_qualifier"]
            for key
            in db_level1_dict
        ]
    )
    qualifier_verbose = get_icf_qualifier('d', '', '.', str(qualifier))
    return {
        'qualifier': qualifier,
        'qualifier_verbose': qualifier_verbose
    }


# andmeid valdkondade jaoks
def get_icf_calcs_prt(request):
    params = request.GET.get('params', '')
    if params:
        codeset = json.loads(params)
    else:
        codeset = []

    db_level1_dict = calc_dblevel1_qualifiers(codeset)
    d_level1_max = calc_dlevel1_max(db_level1_dict)
    db_level1_matrix = render_to_string(
        'main/prt/db_level1_matrix.html',
        context={
            'db_level1_dict': db_level1_dict,
            'd_level1_max': d_level1_max
        }
    )



    b_level2_dict = calc_blevel2_qualifiers(codeset)
    b_level2_matrix = render_to_string(
        'main/prt/b_level2_matrix.html',
        context={'b_level2_dict': b_level2_dict}
    )

    return JsonResponse(
        {
            'db_level1_matrix': db_level1_matrix,
            'b_level2_matrix': b_level2_matrix,
        },
        safe=False
    )

def read_coreset_from_excel():
    wb = openpyxl.load_workbook(settings.BASE_DIR / 'main' / 'static' / 'main' / 'data' / 'RFK_coreset.xlsx')
    ws = wb.active
    # Create a list of functions
    functions = [] 
  
    # Iterate over the rows in the sheet 
    # Iterate through rows 
    for i, row in enumerate(ws): 
        # Skip the first row (the row with the column names) 
        if i == 0: 
            continue
        # Get the value of the first cell in the row 
        fn = row[0].value
        if fn not in functions:
            # Add the value to the list 
            functions.append(fn) 
    
    coreset = {fn: [] for fn in functions}
    
    for col in ws.iter_rows(
            min_row=2, max_row=None,
            min_col=1, max_col=6,
            values_only=True
    ):
        coreset[col[0]].append(
            {
                'valdkond': col[1],
                'kategooria': col[2],
                'vanaduspensioniealine': col[3],
                'kooliealine': col[4],
                'koolieelik': col[5]
            }
        )
    return coreset
    
def coreset(request):
    coreset = read_coreset_from_excel()
    for key in coreset:
        print(key)
        for n in range(0, len(coreset[key])):
            category = coreset[key][n]['kategooria']
            category_row = RFK.objects.get(code=category)
            coreset[key][n] = {
                'valdkond': coreset[key][n]['valdkond'],
                'vanaduspensioniealine': coreset[key][n]['vanaduspensioniealine'],
                'kooliealine': coreset[key][n]['kooliealine'],
                'koolieelik': coreset[key][n]['koolieelik'],
                'category': category_row.code,
                'Translated_title': category_row.Translated_title,
                'Translated_description': category_row.Translated_description,
                'Translated_inclusions': category_row.Translated_inclusions,
                'Translated_exclusions': category_row.Translated_exclusions,
                'form': KategooriaForm(
                    auto_id=f'{key}_%s',
                )
            }
    return render(
        request,
        template_name='main/coreset.html',
        context={
            'coreset': coreset
        }
    )

def expmoodul(request):
    vanusgrupid = ['kõik', 'koolieelik', 'kooliealine', 'tööealine', 'vanaduspensioniealine']
    valitud_vanusgrupp = request.GET.get('vanusgrupp')
    if request.method == "POST":
        form = KategooriaForm(request.POST)
        if form.is_valid():
            kategooria = form.cleaned_data["kategooria"]
            print(kategooria)
    else:
        form = KategooriaForm()
    return render(
        request,
        'main/expmoodul.html',
        {
            'form': form,
            'vanusgrupid': vanusgrupid,
            'valitud_vanusgrupp': valitud_vanusgrupp
        }
    )


def get_helenamiia(request):
    ipaadress = request.META.get('HTTP_X_FORWARDED_FOR') or request.META.get('REMOTE_ADDR')
    return JsonResponse(
        {
            'uussisu': f'<strong>Nüüd on õige asi</strong><br>Sinu aadress on {ipaadress}',
        },
        safe=False
    )

def helenamiia(request):
    metaandmed = request.META
    return render(
        request,
        template_name='main/helenamiia.html',
        context={
            'metaandmed': metaandmed
        }
    )


import os
if __name__ == "__main__":
    import django
    os.environ['DJANGO_SETTINGS_MODULE'] = 'rfk.settings'
    django.setup()

if __name__ == '__main__':
    # for i in range(1, 5):
    #     test(i)
    # get_excel()
    import_icf2db()
    pass
